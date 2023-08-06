import typing

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import Rule
from openpyxl.worksheet.worksheet import Worksheet

from picsexl.excelwriter.exceptions import ExcelWriterException


class ExcelWriter:
    __workbook: typing.Optional[Workbook] = None

    def __init__(
        self,
        workbook_name: str,
        workbook_extension: str = ".xlsx",
        want_cleared: bool = True,
        date_cols: typing.List[str] = None,
        date_format: str = "DD/MM/YYYY HH:MM:SS",
        sheets_to_create: typing.Tuple = (),
        conditional_formatting: typing.List[typing.Tuple[str, Rule]] = None,
    ):
        self.__workbook_name: str = workbook_name
        self.__workbook_extension: str = workbook_extension
        self.__full_workbook_name: str = (
            self.__workbook_name + self.__workbook_extension
        )
        self.__want_cleared: bool = want_cleared
        self.__date_cols: typing.List[str] = date_cols
        self.__date_format: str = date_format
        self.__sheets_to_create: typing.Tuple = sheets_to_create
        self.__conditional_formatting: typing.List[
            typing.Tuple[str, Rule]
        ] = conditional_formatting
        self.__load_or_create_wb()

    def fill_workbook(self, all_data: typing.Dict[str, typing.List]):
        try:
            for k, v in all_data.items():
                self.__create_and_fill_ws(sheet_name=k, data_to_fill=v)
            self.__save_and_close_wb()
        except Exception as exc:
            raise ExcelWriterException(exc=exc)

    @property
    def full_workbook_name(self) -> str:
        return self.__full_workbook_name

    def __load_or_create_wb(self):
        if self.__want_cleared:
            self.__workbook = Workbook()
            self.__workbook.remove(worksheet=self.__workbook.active)

            if self.__sheets_to_create:
                for sc in self.__sheets_to_create:
                    self.__workbook.create_sheet(title=sc)
        else:
            self.__workbook = load_workbook(self.__full_workbook_name, keep_vba=True)

    def __apply_conditional_formatting(self, ws: Worksheet):
        for cf in self.__conditional_formatting:
            ws.conditional_formatting.add(
                range_string=cf[0],
                cfRule=cf[1],
            )

    def __save_and_close_wb(self):
        self.__workbook.save(self.__full_workbook_name)
        self.__workbook.close()

    def __rename_and_pick_first_ws(self, sheet_name: str) -> Worksheet:
        ws: Worksheet = self.__workbook.worksheets[0]
        ws.title = sheet_name
        return ws

    def __create_named_ws_in_wb(self, sheet_name: str) -> Worksheet:
        return self.__workbook.create_sheet(title=sheet_name)

    def __ws_append_with_data(self, ws: Worksheet, data: typing.List):
        if self.__date_cols:
            for c in self.__date_cols:
                ws.column_dimensions[c].number_format = self.__date_format

        for data_row in data:
            ws.append(data_row)

    def __create_and_fill_ws(self, sheet_name: str, data_to_fill: typing.List):
        ws = self.__create_named_ws_in_wb(sheet_name=sheet_name[:30])
        if self.__conditional_formatting:
            self.__apply_conditional_formatting(ws=ws)
        self.__ws_append_with_data(ws=ws, data=data_to_fill)
