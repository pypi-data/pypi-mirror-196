from openpyxl import Workbook
from openpyxl.styles.numbers import NumberFormat
from openpyxl.styles import Alignment
from typing import List, Dict
from ..models import DetailItem, CoreGroup, Section


class ExcelReporter:
    def __init__(self) -> None:
        self.wb = Workbook()

    def auto_width(self, worksheet):
        dims = {}
        for row in worksheet.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))

        for col, value in dims.items():
            worksheet.column_dimensions[col].width = (value + 2) + 2

    def write_title_row(self, sheet, title_row):
        for idx in range(0, len(title_row)):
            cell = sheet.cell(row=1, column=idx + 1)
            cell.value = title_row[idx]

    def write_cell(self, sheet, row, column, value, format = None):
        cell = sheet.cell(row = row, column=column) 
        cell.value = value
        if (format != None):
            if ('alignment' in format):
                cell.alignment = format['alignment']
            if ('number_format' in format):
                cell.number_format = format['number_format']

    def write_detail(self, items: List[DetailItem]):
        sheet = self.wb.create_sheet("Detail", 1)

        title_row = ["Name", "Group", "Core", "Section", "Type", "Size", "StartAddress", "EndAddress"]
        self.write_title_row(sheet, title_row)

        row = 2
        for item in items:
            self.write_cell(sheet, row, 1, item.name)
            self.write_cell(sheet, row, 2, item.group)
            self.write_cell(sheet, row, 3, item.core)
            self.write_cell(sheet, row, 4, item.section)
            self.write_cell(sheet, row, 5, item.type)
            self.write_cell(sheet, row, 6, item.size)
            self.write_cell(sheet, row, 7, "%x" % item.start_addr)
            self.write_cell(sheet, row, 8, "%x" % item.end_addr)
            row += 1

        self.auto_width(sheet)

    def write_component_summary(self, cores: Dict[str, CoreGroup]):
        sheet = self.wb.create_sheet("Summary", 0)
        #sheet = self.wb.active

        title_row = ["Name", "Core", "Files", "TEXT_Used", "RODATA_Used", "DATA_Used", "BSS_Used", "CALIB_Used", "shared", "ROM", "RAM", "Total"]
        self.write_title_row(sheet, title_row)

        row = 2
        for core_name in sorted(cores.keys()):
            for group_name in sorted(cores[core_name].groups.keys()):
                group = cores[core_name].groups[group_name]
                self.write_cell(sheet, row, 1, group_name)
                self.write_cell(sheet, row, 2, core_name)
                #self.write_cell(sheet, row, 2, group.files, Alignment(wrap_text=True, shrink_to_fit=True))
                self.write_cell(sheet, row, 4, group.text_total)
                self.write_cell(sheet, row, 5, group.rodata_total)
                self.write_cell(sheet, row, 6, group.data_total)
                self.write_cell(sheet, row, 7, group.bss_total)
                self.write_cell(sheet, row, 8, group.calib_total)
                self.write_cell(sheet, row, 9, 0)
                self.write_cell(sheet, row, 10, group.total_rom)
                self.write_cell(sheet, row, 11, group.total_ram)
                self.write_cell(sheet, row, 12, group.total)
                row += 1

        self.auto_width(sheet)

    def write_section(self, sections: List[Section]):
        sheet = self.wb.create_sheet("Section", 2)

        title_row = ["Name", "address", "offset", "size", "type"]
        self.write_title_row(sheet, title_row)

        row = 2
        for section in sections:
            self.write_cell(sheet, row, 1, section.name)
            self.write_cell(sheet, row, 2, "%x" % section.base_addr)
            self.write_cell(sheet, row, 3, "%x" % section.offset)
            self.write_cell(sheet, row, 4, section.size)
            self.write_cell(sheet, row, 5, section.type)
            row += 1

        self.auto_width(sheet)

    def write_section_summary(self, sections: List[Section], section_sizes: Dict[str, int]):
        sheet = self.wb.create_sheet("Section Summary", 3)

        total = {}
        total["used"] = 0
        total["size"] = 0

        summary = {}
        for key in section_sizes:
            summary[key] = 0
        for section in sections:
            if not section.type in summary.keys():
                summary[section.type] = 0
            summary[section.type] += section.size
        
        title_row = ["Type", "Used", "Total", "Percent"]
        self.write_title_row(sheet, title_row)

        row = 2
        for key in sorted(summary):
            self.write_cell(sheet, row, 1, key)
            self.write_cell(sheet, row, 2, summary[key])
            total["used"] += summary[key]
            if (key not in section_sizes):
                self.write_cell(sheet, row, 3, "NA", {"alignment": Alignment(horizontal="right")})
                self.write_cell(sheet, row, 4, "NA", {"alignment": Alignment(horizontal="right")})
            else:
                self.write_cell(sheet, row, 3, section_sizes[key])
                self.write_cell(sheet, row, 4, '=B{0}/C{0}'.format(row), {'number_format': "0.00%"})
                total["size"] += section_sizes[key]
            row += 1

        # Write the summary report
        self.write_cell(sheet, row, 1, "")
        self.write_cell(sheet, row, 2, total["used"])
        self.write_cell(sheet, row, 3, total["size"])
        self.write_cell(sheet, row, 4, '=B{0}/C{0}'.format(row), {'number_format': "0.00%"})
            
        
        self.auto_width(sheet)

    def write_core_summary(self):
        pass

    def save(self, name: str):
        self.wb.save(name)
