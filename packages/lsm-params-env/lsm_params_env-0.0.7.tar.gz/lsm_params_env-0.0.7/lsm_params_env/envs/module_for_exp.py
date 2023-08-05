# !/usr/bin/env python3
# _*_ coding:utf-8 _*_
"""
@File     : module_for_exp.py
@Project  : PerfusionCompartmentModel
@Software : PyCharm
@Author   : Xiong YiWei
@Time     : 2022/1/23 14:30
-------------------------------------------------------
                    @Description
    
-------------------------------------------------------
@Last Modify Time      @Version     @Description
2022/1/23 14:30        1.0             None
"""

import json  # 用于json文件读取
import time
from os import listdir

import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import Workbook, load_workbook, styles
from scipy.stats import ttest_ind as ttest
import statistics

from kinetic_modeling import *
# from multi_prior_popul import *


def exp_for_oneCase_prior_multiPopul(directory, fileName, param_l, model, prior_params_data, method, config, **kws):
    """
        对一例肝脏动态PET数据进行先验多种群实验
        Args:
            directory: 数据文件路径
            fileName: 数据文件名（加后缀）
            param_l: list 参数设置列表，包含各参数的下界和上界
                0 ndarray 各参数下界一维数组
                1 ndarray 各参数上界一维数组
            model: kinetic_param_estimate.CompartmentModel 房室模型对象
            method: dict 参数估计优化算法
            prior_params_data: dict
            config: dict 实验设置
            **kws:

        Returns:
            patientName: str 病例患者姓名
            res_normal: dict
            res_tumor: dict
        """
    time_point_start_oneCase = time.perf_counter()  # HACK
    # 导入动态PET数据
    kws_get_dynamicPetData = {}
    if 'split config' in config:
        kws_get_dynamicPetData.update({'split config': config['split config']})
    if 'xlsx_sheetName' in kws:
        kws_get_dynamicPetData.update({'xlsx_sheetName': kws['xlsx_sheetName']})
    patientName, patientData = get_dynamicPetData(directory, fileName, **kws_get_dynamicPetData)
    # 药代动力学房室建模
    modeling = CompartmentModelling(patientName=patientName, data=patientData, model=model,
                                    n_quad=config['num of quadrature order'])
    # region 先验信息实例化
    group_names = tuple(prior_params_data['data'].keys())
    info_type = prior_params_data['type']
    # 将各group键值独立成tuple，并将各动力学参数的先验信息数据的键值独立成tuple
    data_l = copy.deepcopy(list(prior_params_data['data'].values()))  # 每个元素为一组
    for group in data_l:
        for group_key, group_value in zip(group.keys(), group.values()):
            group[group_key] = tuple(group_value.values())
    data = tuple(data_l)
    prior_info = KineticParamPriorInfo(group_names, info_type, *data)
    # endregion

    exp_outOption = ('normal', 'tumor')
    res_normal = None
    res_tumor = None
    for out_option in exp_outOption:  # 分别对该例的normal和tumor输出TAC进行实验
        def residual(params):  # 确定residual函数
            return modeling.compartment_model_residual(params, **{'out_option': out_option, 'out_form': 'RMSE_value'})

        if len(param_l) == 3:
            paramsBounds = param_l[1:3]
        else:
            paramsBounds = param_l
        res = multi_popul(prior_info, residual, paramsBounds, method,
                          config={'classify item': method['others']['multiPopulConfig']['classify item']})
        if out_option == 'normal':
            res_normal = res
        else:
            res_tumor = res

    time_point_end_oneCase = time.perf_counter()  # HACK
    print(f'Spend {(time_point_end_oneCase - time_point_start_oneCase):.4n}s on Case {patientName}!!!')  # HACK

    return patientName, res_normal, res_tumor


def exp_for_oneCase_prior_multiPopul_oneArg(args):
    directory, fileName, param_l, model, method, prior_params_data, config, kws = args
    return exp_for_oneCase_prior_multiPopul(directory, fileName, param_l, model, method, prior_params_data, config,
                                            **kws)


def exp_for_oneCase(directory, fileName, param_l, model, method, config, **kws):
    """
    对一例肝脏动态PET数据进行实验
    Args:
        directory: 数据文件路径
        fileName: 数据文件名（加后缀）
        param_l: list 参数设置列表，包含各参数的起始值、下界和上界
            0 ndarray 各参数起始值一维数组
            1 ndarray 各参数下界一维数组
            3 ndarray 各参数上界一维数组
        model: kinetic_param_estimate.CompartmentModel 房室模型对象
        method: dict 参数估计优化算法
        config: dict 实验设置
        **kws:

    Returns:
        patientName: str 病例患者姓名
        res_normal: curve_fit.curve_fit_result TAC曲线拟合结果数据结构
        res_tumor: curve_fit.curve_fit_result TAC曲线拟合结果数据结构
    """
    time_point_start_oneCase = time.perf_counter()  # HACK
    # 导入动态PET数据
    kws_get_dynamicPetData = {}
    if 'split config' in config:
        kws_get_dynamicPetData.update({'split config': config['split config']})
    if 'xlsx_sheetName' in kws:
        kws_get_dynamicPetData.update({'xlsx_sheetName': kws['xlsx_sheetName']})
    patientName, patientData = get_dynamicPetData(directory, fileName, **kws_get_dynamicPetData)
    modeling = CompartmentModelling(patientName=patientName, data=patientData, model=model)

    res_normal = modeling.model_params_estimate(param_l, out_option='normal', method=method)
    res_tumor = modeling.model_params_estimate(param_l, out_option='tumor', method=method)

    # plot for exp
    if 'plot' in kws:
        if 'interp_compare' in kws['plot']:
            if config['interpConfig']['interp']:
                pass
                # TODO: 预插值处理效果比较绘图
        elif 'fit_compare' in kws['plot']:
            pass
            # TODO: 拟合输出曲线比较绘图
    time_point_end_oneCase = time.perf_counter()  # HACK
    print(f'Spend {(time_point_end_oneCase - time_point_start_oneCase):.4n}s on Case {patientName}!!!')  # HACK
    return patientName, res_normal, res_tumor


def exp_for_oneCase_oneArg(args):
    directory, fileName, param_l, model, method, config, kws = args
    return exp_for_oneCase(directory, fileName, param_l, model, method, config, **kws)


def plotTacCurve(*data, plotMode='standard', **kws):
    if plotMode == 'standard':  # 标准TAC曲线绘图
        # 绘图数据材料
        assert len(data) == 5, 'data中的数据数目不对（应分别为t、artery、vein、normal和tumor的数据array-like）'
        t, artery, vein, normal, tumor = data
        if len(kws) == 0:  # 默认kws
            kws = {
                'figsize': (12.0, 8.0),
                'fmt': ('r-',  # artery 输入曲线 format str
                        'b-',   # vein 输入曲线 format str
                        'g-',   # normal 输出曲线 format str
                        'm-')  # tumor 输出曲线 format str
            }
        # 进行绘图
        if 'broken axis' in kws:
            fig, (ax1, ax2) = plt.subplots(1, 2, sharey='all',
                                           gridspec_kw={'width_ratios': (5, 1), 'height_ratios': (1,)})
            fig.subplots_adjust(hspace=0.02)
            ax1.plot(t, artery, kws['fmt'][0], label='artery')
            ax1.plot(t, vein, kws['fmt'][1], label='vein')
            ax1.plot(t, normal, kws['fmt'][2], label='normal')
            ax1.plot(t, tumor, kws['fmt'][3], label='tumor')
            ax1.set_xlim(-0.5, 5.5)
            ax2.plot(t, artery, kws['fmt'][0], label='artery')  # noqa
            ax2.plot(t, vein, kws['fmt'][1], label='vein')
            ax2.plot(t, normal, kws['fmt'][2], label='normal')
            ax2.plot(t, tumor, kws['fmt'][3], label='tumor')
            ax2.legend()
            ax2.set_xlim(59., 61.)
            plt.show()
        else:
            plt.rcParams['figure.figsize'] = kws['figsize']
            plt.plot(t, artery, kws['fmt'][0], label='artery')  # noqa
            plt.plot(t, vein, kws['fmt'][1], label='vein')
            plt.plot(t, normal, kws['fmt'][2], label='normal')
            plt.plot(t, tumor, kws['fmt'][3], label='tumor')
            plt.legend()
            plt.show()
    elif plotMode == 'fit compare':  # 拟合效果展示绘图
        # 绘图数据材料
        assert len(data) == 5, 'data中的数据数目不对（应分别为t、normal、normal_fit、tumor和tumor_fit的数据array-like）'
        t, normal, normal_fit, tumor, tumor_fit = data
        if len(kws) == 0:  # 默认kws
            kws = {
                'figsize': (12.0, 8.0),
                'fmt': ('g.',  # normal 原输出曲线 format str
                        'g-',   # normal 拟合输出曲线 format str
                        'orange.',   # tumor 原输出曲线 format str
                        'orange-')  # tumor 拟合输出曲线 format str
            }
        # 进行绘图
        plt.rcParams['figure.figsize'] = kws['figsize']
        plt.subplot(1, 2, 1)
        plt.plot(t, normal, kws['fmt'][0], t, normal_fit, kws['fmt'][1])  # normal对照组 原输出:拟合输出
        plt.subplot(1, 2, 2)
        plt.plot(t, tumor, kws['fmt'][2], t, tumor_fit, kws['fmt'][3])  # tumor实验组 原输出:拟合输出
        plt.show()


def get_dynamicPetData(dataFile_directory, dataFile_name, **kws):
    """
    从json文件中提取动态PET数据
    Args:
        dataFile_directory: 数据文件存放路径，仅支持：'.//data//json//60s//'、'.//data//json//5min//'、'.//data//xlsx//5min+60min//'
        dataFile_name: 数据文件名，需给出文件后缀（.json .xlsx）
        **kws:

    Returns:
        name:
        data:
    """
    # 根据路径确定TAC时间数据
    assert dataFile_directory == './/data//json/60s//' or\
           dataFile_directory == './/data//json//5min//' or \
           dataFile_directory == './/data//xlsx//5min+60min//', '不是正确的路径'
    if dataFile_directory == './/data//json//60s//':
        # 60s数据：时间从 1/12min 开始，间隔 1/12min ，均匀增加至 1min 结束
        time_l = np.linspace(1 / 12, 1, 12)
    elif dataFile_directory == './/data//json//5min//':
        # 5min数据：第一段时间从 1/12min 开始，间隔 1/12min ，均匀增加至 1min 结束；
        #          第二段时间从 2min 开始，间隔 1min ，均匀增加至 5min 结束
        time_l = np.concatenate((np.linspace(1 / 12, 1, 12), np.linspace(2, 5, 4)), axis=0)
        # time_l = np.linspace(1 / 12, 16 / 12, 16)
    elif dataFile_directory == './/data//xlsx//5min+60min//':
        # 5min+60min数据：第一段时间从 1/12min 开始，间隔 1/12min ，均匀增加至 1min 结束；
        #                第二段时间从 2min 开始，间隔 1min ，均匀增加至 5min 结束；
        #                第三段时间从 60min 开始且仅此一帧
        time_l = np.concatenate((np.linspace(1 / 12, 1, 12), np.linspace(2, 5, 4), np.array([60])), axis=0)
    else:
        time_l = None

    assert dataFile_name in listdir(dataFile_directory), '路径中的文件不存在'
    assert '.json' in dataFile_name or '.xlsx' in dataFile_name, '不是支持的文件格式'
    name = ''
    data = None
    if dataFile_name in listdir(dataFile_directory) and '.json' in dataFile_name:
        with open(dataFile_directory + dataFile_name) as file:
            jsonData = json.load(file)
        name = jsonData[0]
        if 'split config' in kws:  # 切片
            data = {
                'time': time_l[kws['split config']['start'] - 1 : kws['split config']['end']],
                'concentration': jsonData[1][kws['split config']['start'] - 1 : kws['split config']['end']]
            }
        else:  # 不切片
            data = {
                'time': time_l,
                'concentration': jsonData[1]
            }
    elif dataFile_name in listdir(dataFile_directory) and '.xlsx' in dataFile_name:
        assert 'xlsx_sheetName' in kws, 'kws中没有xlsx_sheetName指定sheet名称，无法确定sheet选项'
        sheet = load_workbook(filename=dataFile_directory + dataFile_name).get_sheet_by_name(kws['xlsx_sheetName'])
        name = kws['xlsx_sheetName']
        data = {
                'time': time_l,
                'concentration': {
                    'artery': np.array([]),
                    'vein': np.array([]),
                    'normal': np.array([]),
                    'tumor': np.array([])
                }
        }
        dict_head_idx_col = {  # 数据表头项列索引字典
            'artery': 0,
            'vein': 0,
            'normal': 0,
            'tumor': 0
        }
        for idx_col in range(1, 7):
            if sheet.cell(row=1, column=idx_col).value in dict_head_idx_col:  # 检测到首行有数据表头项
                dict_head_idx_col[sheet.cell(row=1, column=idx_col).value] = idx_col  # 写入该数据项的列索引
        if 'split config' in kws:  # 切片
            # 先对 time 切片
            data['time'] = data['time'][kws['split config']['start'] - 1 : kws['split config']['end']]
            # 设置sheet的起点和终点行索引，指导后续从excel表的指定sheet导入数据使执行切片
            begin_row = kws['split config']['start'] + 1
            end_row = kws['split config']['end'] + 1
        else:
            begin_row = 2
            end_row = 18
        # 遍历sheet导入TAC数据
        for idx_row in range(begin_row, end_row + 1):  # 不遍历首行表头，从2开始
            data['concentration']['artery'] = np.append(data['concentration']['artery'],
                                                        sheet.cell(row=idx_row,
                                                                   column=dict_head_idx_col['artery']).value)
            data['concentration']['vein'] = np.append(data['concentration']['vein'],
                                                      sheet.cell(row=idx_row,
                                                                 column=dict_head_idx_col['vein']).value)
            data['concentration']['normal'] = np.append(data['concentration']['normal'],
                                                        sheet.cell(row=idx_row,
                                                                   column=dict_head_idx_col['normal']).value)
            data['concentration']['tumor'] = np.append(data['concentration']['tumor'],
                                                       sheet.cell(row=idx_row,
                                                                  column=dict_head_idx_col['tumor']).value)
    return name, data


def export_paramsTables(directory, exp_str, tables):
    xlsWriter = pd.ExcelWriter(directory + exp_str + '.xlsx')
    for tables_key in tables:
        tables[tables_key].to_excel(xlsWriter, sheet_name=tables_key)
    xlsWriter.save()


def export_multiPopulTables(directory, exp_str, tables):
    xlsWriter = pd.ExcelWriter(directory + exp_str + '_multiPopul.xlsx')
    for tables_key in tables:
        tables[tables_key].to_excel(xlsWriter, sheet_name=tables_key)
    xlsWriter.save()


def report_exp_1(directory, exp_str, tables):
    """

    Args:
        directory: str
        exp_str: str
        tables: dict
            'k1', ..., 'f': pd.DataFrame

    Returns:

    """
    book = Workbook()
    book.active.title = 'params & metrics'
    sheet_all = book['params & metrics']
    sheet_stats = book.create_sheet('statistics')
    # region params & metrics 工作表 表头列、行
    names = tables['RMSE']['name'].to_list()
    name_col = ['name'] + names
    for idx, name in enumerate(name_col):
        sheet_all.cell(row=idx + 1, column=1).value = name
        sheet_all.cell(row=idx + 1, column=1).alignment = styles.Alignment(horizontal='center')
        sheet_all.cell(row=idx + 1, column=1).font = styles.Font(bold=True)
    n_name = len(names)
    sheet_all.cell(row=1 + n_name + 1, column=1).value = 'total mean'
    sheet_all.cell(row=1 + n_name + 1, column=1).alignment = styles.Alignment(horizontal='center')
    sheet_all.cell(row=1 + n_name + 1, column=1).font = styles.Font(bold=True)
    sheet_all.cell(row=1 + n_name + 2, column=1).value = 'total std'
    sheet_all.cell(row=1 + n_name + 2, column=1).alignment = styles.Alignment(horizontal='center')
    sheet_all.cell(row=1 + n_name + 2, column=1).font = styles.Font(bold=True)
    sheet_all.cell(row=1 + n_name + 3, column=1).value = 'p value'
    sheet_all.cell(row=1 + n_name + 3, column=1).alignment = styles.Alignment(horizontal='center')
    sheet_all.cell(row=1 + n_name + 3, column=1).font = styles.Font(bold=True)
    for idx, item in zip(range(len(tables)), tables.keys()):
        sheet_all.cell(row=1, column=1 + idx * 2 + 1).value = 'N-' + item
        sheet_all.cell(row=1, column=1 + idx * 2 + 1).alignment = styles.Alignment(horizontal='center')
        sheet_all.cell(row=1, column=1 + idx * 2 + 1).font = styles.Font(bold=True)
        sheet_all.cell(row=1, column=1 + idx * 2 + 2).value = 'T-' + item
        sheet_all.cell(row=1, column=1 + idx * 2 + 2).alignment = styles.Alignment(horizontal='center')
        sheet_all.cell(row=1, column=1 + idx * 2 + 2).font = styles.Font(bold=True)
    # endregion params & metrics 工作表 表头列、行

    # region 写入数据
    for idx, table in enumerate(tables.values()):
        n_col = table['N-value']
        t_col = table['T-value']
        for r in range(len(n_col)):
            sheet_all.cell(row=1 + r + 1, column=1 + idx * 2 + 1).value = n_col[r]
            sheet_all.cell(row=1 + r + 1, column=1 + idx * 2 + 1).alignment = styles.Alignment(horizontal='right')
            sheet_all.cell(row=1 + r + 1, column=1 + idx * 2 + 2).value = t_col[r]
            sheet_all.cell(row=1 + r + 1, column=1 + idx * 2 + 2).alignment = styles.Alignment(horizontal='right')
    # endregion 写入数据

    min_row = sheet_all.min_row
    max_row = sheet_all.max_row - 3  # 除去 total mean、total std 和 p vale
    min_column = sheet_all.min_column
    max_column = sheet_all.max_column
    item_lst = list(tables.keys())
    # region statistics 工作表的表头行和表头列
    sheet_stats.cell(row=1, column=1).value = 'item'
    sheet_stats.cell(row=1, column=2).value = 'N-DIST'
    sheet_stats.cell(row=1, column=3).value = 'T-DIST'
    sheet_stats.cell(row=1, column=4).value = 'p value'
    for c in range(1, 5):
        sheet_stats.cell(row=1, column=c).font = styles.Font(bold=True)
        sheet_stats.cell(row=1, column=c).alignment = styles.Alignment(horizontal='center')
    for idx, key in enumerate(item_lst):
        sheet_stats.cell(row=1 + idx + 1, column=1).value = key
        sheet_stats.cell(row=1 + idx + 1, column=1).font = styles.Font(bold=True)
        sheet_stats.cell(row=1 + idx + 1, column=1).alignment = styles.Alignment(horizontal='center')
    # endregion

    # region 统计均值、标准差和p值
    item_data = {}
    for idx_c in range(min_column + 1, max_column + 1):  # 遍历除了首列的每列
        samp_str = sheet_all.cell(row=1, column=idx_c).value[0]  # 当前列的样本str，'N'或'T'
        item_str = sheet_all.cell(row=1, column=idx_c).value[-2:]  # 当前列的item
        item_str = 'RMSE' if item_str == 'SE' else item_str
        item_data.update({samp_str: []})
        for idx_r in range(min_row + 1, max_row + 1):  # 遍历该列每个包含数据的单元格
            item_data[samp_str].append(sheet_all.cell(row=idx_r, column=idx_c).value)
        mean = statistics.mean(item_data[samp_str])
        sheet_all.cell(row=max_row + 1, column=idx_c).value = mean
        sheet_all.cell(row=max_row + 1, column=idx_c).alignment = styles.Alignment(horizontal='right')
        sheet_all.cell(row=max_row + 1, column=idx_c).number_format = '0.0000'
        sheet_all.cell(row=max_row + 1, column=idx_c).font = styles.Font(color='00DD0000')
        std = statistics.stdev(item_data[samp_str])
        sheet_all.cell(row=max_row + 2, column=idx_c).value = std
        sheet_all.cell(row=max_row + 2, column=idx_c).alignment = styles.Alignment(horizontal='right')
        sheet_all.cell(row=max_row + 2, column=idx_c).number_format = '0.0000'
        sheet_all.cell(row=max_row + 2, column=idx_c).font = styles.Font(color='00DD0000')
        cell_distrib = sheet_stats.cell(row=1 + item_lst.index(item_str) + 1, column=2 if samp_str == 'N' else 3)
        cell_distrib.value = format(mean, '.3f') + '±' + format(std, '.3f')
        cell_distrib.alignment = styles.Alignment(horizontal='right')
        if samp_str == 'T' and item_str != 'RMSE':  # 说明 Normal 和 Tumor 都已经统计完成，可计算p值
            #                  'RMSE'不计算p值
            assert len(item_data['N']) == len(item_data['T']) == max_row - min_row, '已收集的单元格样本数据量不足，无法进行T检验'
            _, p_val = ttest(item_data['N'], item_data['T'])
            sheet_all.cell(row=max_row + 3, column=idx_c - 1).value = p_val
            sheet_all.cell(row=max_row + 3, column=idx_c - 1).number_format = '0.0000'
            sheet_all.merge_cells(start_row=max_row + 3, start_column=idx_c - 1,
                                  end_row=max_row + 3, end_column=idx_c)  # 合并单元格
            sheet_all.cell(row=max_row + 3, column=idx_c - 1).alignment = styles.Alignment(horizontal='center')
            sheet_all.cell(row=max_row + 3, column=idx_c - 1).font = styles.Font(color='00DD0000')
            cell_p = sheet_stats.cell(row=1 + item_lst.index(item_str) + 1, column=4)
            cell_p.value = p_val
            cell_p.number_format = '0.0000'
            cell_p.alignment = styles.Alignment(horizontal='right')
            item_data = {}  # 清空以存储下一个item的数据
    # endregion 统计均值、标准差和p值

    # region 列宽设定
    sheet_all.column_dimensions['A'].width = 16.
    for idx_cols_ord in range(ord('B'), ord('A') + max_column - min_column + 1):
        idx_cols = chr(idx_cols_ord)
        sheet_all.column_dimensions[idx_cols].width = 8.
    sheet_stats.column_dimensions['A'].width = 8.
    sheet_stats.column_dimensions['B'].width = 12.8
    sheet_stats.column_dimensions['C'].width = 12.8
    sheet_stats.column_dimensions['D'].width = 8.
    # endregion 列宽设定

    outFile_name = exp_str + '_REPORT'
    book.save(directory + outFile_name + '.xlsx')


def report_exp_multiPopul(directory, exp_str, tables):
    """
    导出多群体方法的实验结果到excel表格
    原数据tables中应当有各组的params和metrics结果，每个DataFrame的列从name后开始，normal组和tumor组的项交替
    Args:
        directory: str
        exp_str: str
        tables: dict
            '{group name} group {params/metrics}': pd.DataFrame

    Returns:

    """
    # 分类后的结果、normal组的结果和tumor组的结果分别在各自的sheet保存
    report = {'classify': pd.DataFrame([], columns=['params/metrics', 'normal', 'tumor', 'p value']),
              'normal group': pd.DataFrame([], columns=['params/metrics', 'normal', 'tumor', 'p value']),
              'tumor group': pd.DataFrame([], columns=['params/metrics', 'normal', 'tumor', 'p value'])}

    for sheet_key in report:
        # region 确定数据dataFrame
        if sheet_key == 'classify':
            params_df = tables['classified params']
            metrics_df = tables['classified metrics']
        elif sheet_key == 'normal group':
            params_df = tables['normal group params']
            metrics_df = tables['normal group metrics']
        elif sheet_key == 'tumor group':
            params_df = tables['tumor group params']
            metrics_df = tables['tumor group metrics']
        else:
            params_df = None
            metrics_df = None
        # endregion
        idx_df = 0
        # 先统计params
        for i in range(int((len(params_df.columns) - 1) / 2)):  # 遍历所有param
            colName_normal = params_df.columns.values[i * 2 + 1]  # noqa
            colName_tumor = params_df.columns.values[i * 2 + 2]
            param_name = params_df.columns.values[i * 2 + 1].replace('normal ', '')
            param_normal = params_df[colName_normal].values
            param_tumor = params_df[colName_tumor].values
            # 计算对照组分布
            mean_normal = param_normal.mean()
            std_normal = param_normal.std()
            distribution_normal = format(mean_normal, '.2f') + '+/-' + format(std_normal, '.2f')
            # 计算病灶组分布
            mean_tumor = param_tumor.mean()
            std_tumor = param_tumor.std()
            distribution_tumor = format(mean_tumor, '.2f') + '+/-' + format(std_tumor, '.2f')
            # 统计分析（独立样本T检验）
            _, p_value = ttest(param_normal, param_tumor)

            report[sheet_key].loc[idx_df] = [param_name, distribution_normal, distribution_tumor, p_value]
            idx_df += 1
        # 再统计metrics
        for i in range(int((len(metrics_df.columns) - 1) / 2)):  # 遍历所有metrics（i为metric的序号）
            colName_normal = metrics_df.columns.values[i * 2 + 1]  # noqa
            colName_tumor = metrics_df.columns.values[i * 2 + 2]
            metric_name = metrics_df.columns.values[i * 2 + 1].replace('normal ', '')
            metric_normal = metrics_df[colName_normal].values
            metric_tumor = metrics_df[colName_tumor].values
            # 计算对照组分布
            mean_normal = metric_normal.mean()
            std_normal = metric_normal.std()
            distribution_normal = format(mean_normal, '.2f') + '+/-' + format(std_normal, '.2f')
            # 计算病灶组分布
            mean_tumor = metric_tumor.mean()
            std_tumor = metric_tumor.std()
            distribution_tumor = format(mean_tumor, '.2f') + '+/-' + format(std_tumor, '.2f')

            report[sheet_key].loc[idx_df] = [metric_name, distribution_normal, distribution_tumor, 'no p-value']
            idx_df += 1

    outFile_name = exp_str + '_multiPopul_REPORT'
    xlsWriter = pd.ExcelWriter(directory + outFile_name + '.xlsx')
    for sheet_key in report:
        report[sheet_key].to_excel(xlsWriter, sheet_name=sheet_key)
    xlsWriter.save()
