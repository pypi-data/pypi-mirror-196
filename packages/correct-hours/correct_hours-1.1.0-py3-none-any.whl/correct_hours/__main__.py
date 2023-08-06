from openpyxl import load_workbook
from argparse import ArgumentParser
from pathlib import Path

from correct_hours.report_processors.myob import MyobReportProcessor
from correct_hours.report_processors.xero import XeroReportProcessor
from correct_hours.types import RateFileNotFound, InvalidReportType, RATES_FILENAME, OUTPUT_FOLDER, \
    HOURS_NEW_FILE_PREFIX

parser = ArgumentParser()
parser.add_argument("directory", help="Location of Excel files", type=str)
parser.add_argument(
    "-t",
    "--report-type",
    dest="report_type",
    help="Report type",
    type=str,
    default="xero",
    choices=['xero', 'myob']
)

args = parser.parse_args()
directory = args.directory
report_type = args.report_type


def get_new_file_name(filepath: Path) -> str:
    path = Path(filepath)
    # Example: ~/correct-hours/examples/xero/output/copy_barney-stinson.xlsx
    return f"{path.parent.absolute()}/{OUTPUT_FOLDER}/{HOURS_NEW_FILE_PREFIX}{path.name}"


def should_ignore_file(file: Path) -> bool:
    filename = file.name
    return (
        str.startswith(filename, "~") or
        filename == RATES_FILENAME
    )


# create output folder
Path(f"{directory}/{OUTPUT_FOLDER}").mkdir(parents=True, exist_ok=True)
# iterate through files in input directory
files = Path(directory).glob('*')
for f in files:
    if f.is_file():
        if not should_ignore_file(f):
            hours_filepath = f.absolute()
            print(f"Processing file {hours_filepath}...")
            workbook = load_workbook(filename=hours_filepath)
            # load rates workbook
            rates_filepath = f"{directory}/{RATES_FILENAME}"
            if report_type == 'xero':
                try:
                    rates_workbook = load_workbook(filename=rates_filepath)
                except FileNotFoundError:
                    raise RateFileNotFound(rates_filepath)
                processor = XeroReportProcessor(workbook, rates_workbook)
            elif report_type == 'myob':
                processor = MyobReportProcessor(workbook)
            else:
                raise InvalidReportType(report_type)
            processor.process()
            hours_new_filename = get_new_file_name(hours_filepath)
            workbook.save(filename=hours_new_filename)
            print(f"Finished processing file. Created file {hours_new_filename}.")
