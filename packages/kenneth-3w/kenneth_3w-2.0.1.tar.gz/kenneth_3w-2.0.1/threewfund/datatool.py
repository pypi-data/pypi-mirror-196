import re
import time
import datetime
import requests
import numpy as np
import  pandas as pd
import openpyxl.utils.cell as c
from bs4 import BeautifulSoup
from selenium import webdriver
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from pandas import DataFrame,Series
from win32com.client import Dispatch
from openpyxl.formula.translate import Translator


def find_income(company_id, n, google_driver_path, com_revenue_monthly_path, date_col, rev_col):
    driver = webdriver.Chrome(executable_path=google_driver_path)
    driver.get("https://mops.twse.com.tw/mops/web/t05st10_ifrs")
    time.sleep(1)
    input_from_list = driver.find_element_by_id("co_id").send_keys(str(company_id))
    time.sleep(1)
    button = driver.find_element_by_css_selector('[value=" 查詢 "]')
    button.click()
    windows = driver.window_handles
    driver.switch_to.window(windows[-1])
    time.sleep(1.5)
    html = driver.page_source
    soup = BeautifulSoup(html, "html.parser")
    driver.close()
    table = soup.find_all("div", {"id": "table01"})
    for t in table:
        tr_tag = t.find_all('tr')
    if tr_tag == []:
        find_income(company_id, n, google_driver_path, com_revenue_monthly_path, date_col, rev_col)
    else:
        new_income = tr_tag[4].text[3:].split()
        new_date = tr_tag[2].text[:9].split()
        if new_date[0][:2] == '民國':
            new_date = new_date
        else:
            new_date = tr_tag[1].text[:9].split()
        wb = load_workbook(filename=com_revenue_monthly_path)
        ws = wb.active
        ws._current_row = n
        ws.append({rev_col: new_income[0]})
        ws._current_row = n
        ws.append({date_col: new_date[0]})
        wb.save(filename=com_revenue_monthly_path)


def find_row(company_name, sheetname):
    company_row = 'None'
    for col in sheetname.columns:
        for row in col:
            if row.value == company_name:
                company_row = str(row.row)
    return company_row


def find_col(date, sheetname):
    for row in sheetname.rows:
        for col in row:
            if col.value == date:
                date_col = c.get_column_letter(col.col_idx)
                return date_col


def find_address(com_name, date, sheetname):
    if find_col(date, sheetname) != 'None' and find_row(com_name, sheetname) != 'None':
        address = find_col(date, sheetname) + find_row(com_name, sheetname)
    else:
        address = "None"
    return address


def just_open(filename):
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    xlBook = xlApp.Workbooks.Open(filename)
    xlBook.Save()
    xlBook.Close()

def twweb_to_excel(num,google_driver_path,com_revenue_monthly_path,date_col, rev_col):
    wb = load_workbook(com_revenue_monthly_path)
    ws = wb.active
    v = []
    for row in ws.rows:
        v.append(row[0].value)
    for i in range(num,182):
        find_income(v[i][:4], i, google_driver_path,com_revenue_monthly_path,date_col, rev_col)
        print(i)

def find_cunchu_price(chrome_path):
    driver_cunchu = webdriver.Chrome(executable_path=chrome_path)
    driver_cunchu.get('https://www.dramx.com/')
    html = driver_cunchu.page_source
    soup = BeautifulSoup(html,"html.parser")
    table_time = soup.find_all("div",{"class":'menudiv'})
    for t in table_time:
        td_tag = t.find_all('time')
        dram_flash_time = str(td_tag[0])[6:16]
    table_dram = soup.find_all("div",{"id":'con_one_1'})
    for t in table_dram:
        td_tag = t.find_all('td')
    dram_price_a = [td_tag[5],td_tag[12],td_tag[19],td_tag[26],td_tag[33],td_tag[40]]
    dram_price_change_a = [td_tag[6],td_tag[13],td_tag[20],td_tag[27],td_tag[34],td_tag[41]]
    dram_price = []
    dram_price_change = []
    for i in dram_price_a:
        result = float(str(i)[4:9])
        dram_price.append(result)
    for i in dram_price_change_a:
        a = str(i)[-11:-5]
        if a[0] =='>':
            result = float(a[1:-1])/100
        else:
            result = 0 - float(a[1:-1])/100
        dram_price_change.append(result)
    table_flash = soup.find_all("div",{"id":'con_one_2'})
    for t in table_flash:
        td_tag = t.find_all('td')
    flash_price_a = [td_tag[5],td_tag[12],td_tag[19],td_tag[26],td_tag[33]]
    flash_price_change_a = [td_tag[6],td_tag[13],td_tag[20],td_tag[27],td_tag[34]]
    flash_price = []
    flash_price_change = []
    for i in flash_price_a:
        result = float(str(i)[4:9])
        flash_price.append(result)
    for i in flash_price_change_a:
        a = str(i)[-11:-5]
        if a[0] =='>':
            result = float(a[1:-1])/100
        else:
            result = 0 - float(a[1:-1])/100
        flash_price_change.append(result)
    driver_cunchu.close()
    driver_cunchu = webdriver.Chrome("C:\Program Files\Google\Chrome\Application\chromedriver.exe")
    driver_cunchu.get('https://www.0101ssd.com/c/price_new.html')
    html = driver_cunchu.page_source
    soup = BeautifulSoup(html,"html.parser")
    table = soup.find_all("div",{"class":'quotedpriceLeft wid_700'})
    for t in table:
        td_tag = t.find_all('td')
    ssd_price_a = [td_tag[2],td_tag[9],td_tag[16],td_tag[23],td_tag[30],
                  td_tag[37],td_tag[44],td_tag[51],td_tag[72],td_tag[79],
                  td_tag[86],td_tag[93],td_tag[100],td_tag[107]]
    ssd_price_change_a = [td_tag[5],td_tag[12],td_tag[19],td_tag[26],td_tag[33],
                  td_tag[40],td_tag[47],td_tag[54],td_tag[75],td_tag[82],
                  td_tag[89],td_tag[96],td_tag[103],td_tag[110]]
    ssd_price = []
    ssd_price_change = []
    for i in ssd_price_a:
        a = int(str(i)[4:-5])
        ssd_price.append(a)
    for i in ssd_price_change_a:
        x = str(i)
        if x.find('-') == -1:
            n = x.find('>')
            result = float(x[n+1:-6])/100
        else:
            if x[4:-5] == '-':
                result = '-'
            else:
                n = x.find('-')
                result = 0 - float(x[n+1:-6])/100
        ssd_price_change.append(result)
    table = soup.find_all("b",{"class":'quotedTableOneNameTime'})
    for t in table:
        td_tag = t.find_all('td')
    table = soup.find_all("b",{"class":'quotedTableOneNameTime'})
    ssd_time= str(table[0])[-19:-4]
    driver_cunchu.close()
    print('dram和flash更新时间：' + dram_flash_time)
    print('dram价格和涨跌')
    print(dram_price)
    print(dram_price_change)
    print('flash价格和涨跌')
    print(flash_price)
    print(flash_price_change)
    print('ssd' + ssd_time)
    print(ssd_price)
    print(ssd_price_change)
    return dram_price, dram_price_change, flash_price, flash_price_change, ssd_price, ssd_price_change

def jd_start():
    driver.get("https://item.jd.com/100018807455.html")

def jd_find_price_stock(web):
    if web == 'https://jd.com/':
        result = ['暂无信息', '']
    else:
        product_num = web[20:-5]
        price_code = 'price J-p-' + product_num
        driver.get(web)
        html = driver.page_source
        soup = BeautifulSoup(html,"html.parser")
        table_price = soup.find_all("span",{"class":price_code})
        if table_price == []:
            result = ['已下架', '']
        else:
            price_str = str(table_price[0])
            price_location = 25 + len(product_num)
            price = int(price_str[price_location:-10])
            table_stock = soup.find_all("div",{"class":'store-prompt'})
            stock_str = str(table_stock[0])
            stock = stock_str[52:-23]
            if stock == '无货':
                result = [price, stock]
            else:
                result = [price, '有货']
    return result

def find_auto_data(web,year, month, google_driver_address):
    driver = webdriver.Chrome(executable_path=google_driver_address)
    driver.get(web)
    html = driver.page_source
    soup = BeautifulSoup(html,"html.parser")
    table = soup.find_all("div",{"id":'daisuFlash'})
    for t in table:
        td_tag = t.find_all('script')
    a = str(td_tag[2])[15:-10].split(',')
    b = str(td_tag[2])[27:-10].split(',')
    if a[0][0:8] == 'Data'+ year:
        x = b[month - 1][-9:]
        if x == 'undefined':
            result = x
        else:
            result = int(re.sub(u"([^\u0030-\u0039])", "", b[month - 1][-10:]))
    else:
        result = 'no data'
    driver.close()
    return result